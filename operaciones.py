# operaciones.py

import os
import pandas as pd
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from datetime import datetime

# ----------- DEFINIR EL MAPEO DE COLUMNAS GLOBALMENTE -----------
# Definir el mapeo de nombres de columnas
COLUMN_MAPPING = {
    'Bruto': 'Peso bruto (kg)',
    'Neto': 'Peso neto (kg)',
    'Volumen': 'Volumen (m3)',
    'Importe': 'Valor FOB',
    'Doc.comer.': 'Pedido',
    'LibrUtiliz': 'Cajas',
    'Total peso Bruto': 'Total peso Bruto (kg)',
    'Total peso Neto': 'Total peso Neto (kg)',
    'Total Volumen': 'Total volumen',
    'Total Importe': 'Ventas totales FOB',
    'Total LibrUtiliz': 'Total Cajas'
}

def exportar_a_excel(contenedores_volumenes, mensajes_contenedores, df_final, columnas_mapeadas):
    try:
        # Obtener el 'Nombre' al que se le hizo el procedimiento
        nombre_procedimiento = df_final['Nombre'].iloc[0] if 'Nombre' in df_final.columns else 'exportado'
        nombre_procedimiento = ''.join(c for c in nombre_procedimiento if c.isalnum() or c in (' ', '.', '_')).rstrip()

        # Obtener la fecha y hora actuales
        fecha_hora_actual = datetime.now().strftime('%Y-%M-%d %H-%M-%S')

        # Construir el nombre del archivo
        base_filename = f"{nombre_procedimiento} {fecha_hora_actual}.xlsx"

        # Determinar la ruta del escritorio según el sistema operativo
        if os.name == 'nt':  # Para Windows
            desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
        else:  # Para macOS y Linux
            desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')

        filepath = os.path.join(desktop_path, base_filename)

        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            for i, (volumen_contenedor, mensajes) in enumerate(zip(contenedores_volumenes, mensajes_contenedores)):
                df_contenedor = df_final.loc[mensajes].copy()

                df_contenedor.rename(columns={
                    'Bruto': 'Peso bruto (kg)',
                    'Neto': 'Peso neto (kg)',
                    'Volumen': 'Volumen (m3)',
                    'Importe': 'Valor FOB',
                    'Doc.comer.': 'Pedido',
                    'LibrUtiliz': 'Cajas'
                }, inplace=True)

                # Asegurarse de que la columna 'Omitido' existe
                if 'Omitido' not in df_contenedor.columns:
                    df_contenedor['Omitido'] = ''

                # Agregar la columna 'Volumen Cajas'
                df_contenedor['Volumen Cajas'] = ''

                # Reordenar las columnas sin incluir nada en la columna O y moviendo 'Volumen Cajas' a U (columna 21)
                columnas_ordenadas = ['Material', 'Texto de mensaje', 'Cajas', 'Peso bruto (kg)', 'Peso neto (kg)',
                                      'Volumen (m3)', 'Valor FOB', 'Cliente', 'Nombre', 'Contador', 'Pedido',
                                      'Grupo', 'Lote', 'Omitido', 'Volumen Cajas']

                # Verificar que todas las columnas existen en df_contenedor
                columnas_existentes = [col for col in columnas_ordenadas if col in df_contenedor.columns]
                df_contenedor = df_contenedor[columnas_existentes]

                hoja_nombre = f'Contenedor_{i+1}'
                df_contenedor.to_excel(writer, sheet_name=hoja_nombre, index=False)

                workbook = writer.book
                worksheet = writer.sheets[hoja_nombre]

                worksheet.sheet_view.showGridLines = False

                title_fill = PatternFill(start_color="001955", end_color="001955", fill_type="solid")
                title_font = Font(color="FFFFFF", bold=True, size=14)
                title_alignment = Alignment(horizontal="center", vertical="center")

                # Aplicar los estilos a todos los encabezados
                for col in range(1, len(df_contenedor.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.fill = title_fill
                    cell.font = title_font
                    cell.alignment = title_alignment

                # Agregar encabezado para la columna R (Bruto X Cajas)
                worksheet.cell(row=1, column=18).value = "Bruto X Cajas"
                worksheet.cell(row=1, column=18).fill = title_fill
                worksheet.cell(row=1, column=18).font = title_font
                worksheet.cell(row=1, column=18).alignment = title_alignment

                # Agregar la columna "Bruto x Cajas" en la columna R (columna 18)
                bruto_col_idx = df_contenedor.columns.get_loc('Peso bruto (kg)') + 1
                cajas_col_idx = df_contenedor.columns.get_loc('Cajas') + 1

                for row in range(2, 101):  # Filas de 2 a 100
                    bruto_cell_ref = worksheet.cell(row=row, column=bruto_col_idx).coordinate
                    cajas_cell_ref = worksheet.cell(row=row, column=cajas_col_idx).coordinate
                    formula = f'=IF(AND(ISNUMBER({bruto_cell_ref}), ISNUMBER({cajas_cell_ref})), {bruto_cell_ref}*{cajas_cell_ref}, "")'
                    worksheet.cell(row=row, column=18).value = formula  # Columna 18 es R (Bruto x Cajas)

                # Agregar encabezado para la columna S (Fob X Cajas)
                worksheet.cell(row=1, column=19).value = "Fob X Cajas"
                worksheet.cell(row=1, column=19).fill = title_fill
                worksheet.cell(row=1, column=19).font = title_font
                worksheet.cell(row=1, column=19).alignment = title_alignment

                # Agregar la fórmula en la columna S (columna 19) para Fob X Cajas
                valor_fob_col_idx = df_contenedor.columns.get_loc('Valor FOB') + 1
                contador_col_idx = df_contenedor.columns.get_loc('Contador') + 1

                for row in range(2, 101):  # Filas de 2 a 100
                    valor_fob_cell_ref = worksheet.cell(row=row, column=valor_fob_col_idx).coordinate
                    contador_cell_ref = worksheet.cell(row=row, column=contador_col_idx).coordinate
                    cajas_cell_ref = worksheet.cell(row=row, column=cajas_col_idx).coordinate
                    formula_fob = f'=IF(AND(ISNUMBER({valor_fob_cell_ref}), ISNUMBER({contador_cell_ref}), ISNUMBER({cajas_cell_ref})), {valor_fob_cell_ref}*{contador_cell_ref}*{cajas_cell_ref}, "")'
                    worksheet.cell(row=row, column=19).value = formula_fob  # Columna 19 es S (Fob x Cajas)

                # Agregar encabezado para la columna T (Neto X Cajas)
                worksheet.cell(row=1, column=20).value = "Neto X Cajas"
                worksheet.cell(row=1, column=20).fill = title_fill
                worksheet.cell(row=1, column=20).font = title_font
                worksheet.cell(row=1, column=20).alignment = title_alignment

                # Agregar la fórmula en la columna T (columna 20) para Neto X Cajas
                neto_col_idx = df_contenedor.columns.get_loc('Peso neto (kg)') + 1

                for row in range(2, 101):  # Filas de 2 a 100
                    neto_cell_ref = worksheet.cell(row=row, column=neto_col_idx).coordinate
                    cajas_cell_ref = worksheet.cell(row=row, column=cajas_col_idx).coordinate
                    formula_neto = f'=IF(AND(ISNUMBER({neto_cell_ref}), ISNUMBER({cajas_cell_ref})), {neto_cell_ref}*{cajas_cell_ref}, "")'
                    worksheet.cell(row=row, column=20).value = formula_neto  # Columna 20 es T (Neto x Cajas)

                # Mover la columna 'Volumen Cajas' a la columna U (21)
                volumen_col_idx = df_contenedor.columns.get_loc('Volumen (m3)') + 1  # Inicialización de volumen_col_idx
                worksheet.cell(row=1, column=21).value = "Volumen Cajas"  # Columna 21 es U
                worksheet.cell(row=1, column=21).fill = title_fill  # Aplicar el color al encabezado
                worksheet.cell(row=1, column=21).font = title_font  # Aplicar la fuente
                worksheet.cell(row=1, column=21).alignment = title_alignment  # Aplicar la alineación

                for row in range(2, 101):  # Ajusta según tu cantidad de filas
                    cajas_cell_ref = worksheet.cell(row=row, column=cajas_col_idx).coordinate
                    volumen_cell_ref = worksheet.cell(row=row, column=volumen_col_idx).coordinate
                    worksheet.cell(row=row, column=21).value = f'=IF(AND(ISNUMBER({cajas_cell_ref}), ISNUMBER({volumen_cell_ref})), {cajas_cell_ref}*{volumen_cell_ref}, "")'  # Columna 21 es U

                # Dejar la columna O vacía
                worksheet.cell(row=1, column=15).value = ""  # Columna 15 es O, título vacío
                for row in range(2, 101):  # Filas de datos
                    worksheet.cell(row=row, column=15).value = ""  # Columna O, datos vacíos

                total_label_col = 'P'  # Columna 16 (P)
                total_value_col = 'Q'  # Columna 17 (Q)

                bold_font_total = Font(bold=True, size=14)
                alignment_total = Alignment(horizontal="center", vertical="center")

                total_names = [
                    'Total peso Bruto (kg)',
                    'Total peso Neto (kg)',
                    'Total volumen',
                    'Ventas totales FOB',
                    'Total Cajas'
                ]
                totals_start_row = df_contenedor.shape[0] + 2

                for idx, total_name in enumerate(total_names):
                    cell_row = totals_start_row + idx
                    label_cell = worksheet[f'{total_label_col}{cell_row}']
                    value_cell = worksheet[f'{total_value_col}{cell_row}']

                    label_cell.value = total_name
                    label_cell.font = bold_font_total
                    label_cell.alignment = alignment_total

                    end_row = 100  # Ajustar el número de filas

                    omitido_col_idx = df_contenedor.columns.get_loc('Omitido') + 1
                    cajas_col_idx = df_contenedor.columns.get_loc('Cajas') + 1

                    if total_name == 'Total peso Bruto (kg)':
                        bruto_col_idx = df_contenedor.columns.get_loc('Peso bruto (kg)') + 1
                        formula = f"=SUMPRODUCT((LOWER(${get_column_letter(omitido_col_idx)}$2:${get_column_letter(omitido_col_idx)}${end_row})<>\"x\")*(${get_column_letter(bruto_col_idx)}$2:${get_column_letter(bruto_col_idx)}${end_row})*(${get_column_letter(cajas_col_idx)}$2:${get_column_letter(cajas_col_idx)}${end_row}))"
                    elif total_name == 'Total peso Neto (kg)':
                        neto_col_idx = df_contenedor.columns.get_loc('Peso neto (kg)') + 1
                        formula = f"=SUMPRODUCT((LOWER(${get_column_letter(omitido_col_idx)}$2:${get_column_letter(omitido_col_idx)}${end_row})<>\"x\")*(${get_column_letter(neto_col_idx)}$2:${get_column_letter(neto_col_idx)}${end_row})*(${get_column_letter(cajas_col_idx)}$2:${get_column_letter(cajas_col_idx)}${end_row}))"
                    elif total_name == 'Total volumen':
                        volumen_col_idx = df_contenedor.columns.get_loc('Volumen (m3)') + 1
                        formula = f"=SUMPRODUCT((LOWER(${get_column_letter(omitido_col_idx)}$2:${get_column_letter(omitido_col_idx)}${end_row})<>\"x\")*(${get_column_letter(volumen_col_idx)}$2:${get_column_letter(volumen_col_idx)}${end_row})*(${get_column_letter(cajas_col_idx)}$2:${get_column_letter(cajas_col_idx)}${end_row}))"
                    elif total_name == 'Ventas totales FOB':
                        valor_fob_col_idx = df_contenedor.columns.get_loc('Valor FOB') + 1
                        contador_col_idx = df_contenedor.columns.get_loc('Contador') + 1
                        formula = f"=SUMPRODUCT((LOWER(${get_column_letter(omitido_col_idx)}$2:${get_column_letter(omitido_col_idx)}${end_row})<>\"x\")*(${get_column_letter(valor_fob_col_idx)}$2:${get_column_letter(valor_fob_col_idx)}${end_row})*(${get_column_letter(contador_col_idx)}$2:${get_column_letter(contador_col_idx)}${end_row})*(${get_column_letter(cajas_col_idx)}$2:${get_column_letter(cajas_col_idx)}${end_row}))"
                    elif total_name == 'Total Cajas':
                        formula = f"=SUMPRODUCT((LOWER(${get_column_letter(omitido_col_idx)}$2:${get_column_letter(omitido_col_idx)}${end_row})<>\"x\")*(${get_column_letter(cajas_col_idx)}$2:${get_column_letter(cajas_col_idx)}${end_row}))"

                    value_cell.value = formula
                    value_cell.font = bold_font_total
                    value_cell.alignment = alignment_total

                merged_row = totals_start_row + len(total_names) + 2  # Dos filas debajo de los totales
                worksheet.merge_cells(start_row=merged_row, start_column=1, end_row=merged_row, end_column=13)
                merged_cell = worksheet.cell(row=merged_row, column=1)
                merged_cell.value = 'Adición De Referencias'
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                merged_cell.font = Font(bold=True, size=14)

                violet_fill = PatternFill(start_color="7696ff", end_color="7696ff", fill_type="solid")
                for col in range(1, 14):  # Columnas A (1) a M (13)
                    cell = worksheet.cell(row=merged_row, column=col)
                    cell.fill = violet_fill

                alignment_wrap = Alignment(wrap_text=True)
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                            cell.alignment = alignment_wrap
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    if adjusted_width > 15:
                        adjusted_width = 15  # Reducimos el ancho máximo para que quepan más columnas
                    worksheet.column_dimensions[column].width = adjusted_width

                worksheet.sheet_view.zoomScale = 90
                worksheet.sheet_view.topLeftCell = 'A1'

        filas_exportadas = sum(len(mensajes) for mensajes in mensajes_contenedores)
        total_filas = df_final.shape[0]
        if filas_exportadas != total_filas:
            print(f"⚠️ Error de Exportación: Se han exportado {filas_exportadas} filas, pero el total es {total_filas}.")
        else:
            print(f"✅ Todas las filas han sido exportadas correctamente a Excel: {base_filename}")

        messagebox.showinfo("Exportación Completa", f"El archivo ha sido exportado exitosamente:\n{filepath}")

    except Exception as e:
        print(f"Error durante la exportación: {str(e)}")



def calcular_totales(df):
    # Asegurar que las columnas sean numéricas
    for col in ['Bruto', 'Neto', 'Volumen', 'Importe', 'LibrUtiliz', 'Contador']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    # Calcular totales multiplicando por 'LibrUtiliz' donde corresponda
    total_bruto = round((df['Bruto'] * df['LibrUtiliz']).sum(), 2)
    total_neto = round((df['Neto'] * df['LibrUtiliz']).sum(), 2)
    total_volumen = round((df['Volumen'] * df['LibrUtiliz']).sum(), 2)
    total_importe = round((df['Importe'] * df['Contador'] * df['LibrUtiliz']).sum(), 2)
    total_librUtiliz = round(df['LibrUtiliz'].sum(), 2)

    return {
        'Total peso Bruto': total_bruto,
        'Total peso Neto': total_neto,
        'Total Volumen': total_volumen,
        'Total Importe': total_importe,
        'Total LibrUtiliz': total_librUtiliz
    }

def calcular_contenedores(df, columnas_mapeadas, capacidad_contenedor_max):
    # Crear una nueva columna 'Volumen_LibrUtiliz'
    df['Volumen_LibrUtiliz'] = df['Volumen'] * df['LibrUtiliz']
    df['Color_Volumen'] = ''
    
    # Ordenar el DataFrame por 'Lote' de menor a mayor
    df['Lote'] = df['Lote'].astype(str).str.zfill(10)  # Zfill para facilitar la ordenación numérica
    df = df.sort_values(by='Lote').reset_index()

    # Inicializar listas para contenedores
    contenedores_volumenes = []
    mensajes_contenedores = []
    container_volumes = []

    total_rows = df.shape[0]
    processed_rows = 0

    for idx, row in df.iterrows():
        processed_rows += 1
        vol_lib_util = row['Volumen_LibrUtiliz']
        mensaje = row['index']  # Índice original antes del reset

        # Agregar GZ1A si pertenece al grupo GCFOODS
        if row['Grupo'] == 'GCFOODS':
            row['Lote'] = f"{row['Lote']}GZ1A"

        # Si el volumen del mensaje excede la capacidad, asignarlo a su propio contenedor
        if vol_lib_util > capacidad_contenedor_max:
            df.at[idx, 'Color_Volumen'] = 'yellow'
            contenedores_volumenes.append(vol_lib_util)
            mensajes_contenedores.append([mensaje])
            container_volumes.append(vol_lib_util)
            print(f"Fila {processed_rows}/{total_rows} asignada a contenedor individual por exceso de volumen.")
            continue

        # Intentar colocar el elemento en el primer contenedor donde quepa
        placed = False
        for i in range(len(container_volumes)):
            if container_volumes[i] + vol_lib_util <= capacidad_contenedor_max:
                container_volumes[i] += vol_lib_util
                mensajes_contenedores[i].append(mensaje)
                placed = True
                print(f"Fila {processed_rows}/{total_rows} agregada al Contenedor {i+1}. Volumen acumulado: {container_volumes[i]:.2f}")
                break
        if not placed:
            # Crear un nuevo contenedor
            container_volumes.append(vol_lib_util)
            mensajes_contenedores.append([mensaje])
            print(f"Fila {processed_rows}/{total_rows} iniciando un nuevo Contenedor. Volumen: {vol_lib_util:.2f}")

    contenedores_volumenes = container_volumes

    print(f"Total de filas procesadas: {processed_rows}/{total_rows}")
    print(f"Total de contenedores creados: {len(contenedores_volumenes)}")

    # Verificar que todas las filas hayan sido asignadas
    filas_asignadas = sum(len(mensajes) for mensajes in mensajes_contenedores)
    total_filas = df.shape[0]
    if filas_asignadas != total_filas:
        print(f"⚠️ Error: Se han asignado {filas_asignadas} filas, pero el total es {total_filas}.")
    else:
        print("✅ Todas las filas han sido asignadas correctamente a contenedores.")

    return contenedores_volumenes, mensajes_contenedores



def mostrar_resultados(totales, contenedores_volumenes, mensajes_contenedores, df_final, columnas_mapeadas, capacidad_contenedor_max):
    root = tk.Toplevel()
    root.title("Resultados de Contenedores")
    root.geometry("900x600")  # Ajuste del tamaño de la ventana

    # Configurar estilo
    style = ttk.Style(root)
    style.theme_use("clam")  # Puedes cambiar el tema según preferencia

    # Definir estilos personalizados
    style.configure("Treeview.Heading", font=("Arial", 12, "bold"), background="#001955", foreground="white")
    style.configure("Treeview", font=("Arial", 12), rowheight=30, fieldbackground="#f0f0f0")
    style.map("Treeview", background=[("selected", "#ADD8E6")], foreground=[("selected", "black")])

    # Frame para Totales
    frame_totales = ttk.LabelFrame(root, text="Totales", padding=(20, 10))
    frame_totales.pack(fill='x', padx=20, pady=10)

    # Treeview para Totales
    columnas_totales = ("Descripción", "Valor")
    tree_totales = ttk.Treeview(frame_totales, columns=columnas_totales, show="headings", height=5)
    tree_totales.heading("Descripción", text="Descripción")
    tree_totales.heading("Valor", text="Valor")
    tree_totales.column("Descripción", anchor='w', width=300)
    tree_totales.column("Valor", anchor='center', width=300)

    # Insertar datos de totales
    for desc, valor in totales.items():
        # Usar el mapeo para los nombres de totales
        desc_mapeado = COLUMN_MAPPING.get(desc, desc)
        tree_totales.insert("", "end", values=(desc_mapeado, f"{valor:,.2f}"))

    # Scrollbar para Totales
    scrollbar_totales = ttk.Scrollbar(frame_totales, orient="vertical", command=tree_totales.yview)
    tree_totales.configure(yscroll=scrollbar_totales.set)
    scrollbar_totales.pack(side='right', fill='y')
    tree_totales.pack(fill='x')

    # Frame para Contenedores
    frame_contenedores = ttk.LabelFrame(root, text="Contenedores", padding=(20, 10))
    frame_contenedores.pack(fill='both', expand=True, padx=20, pady=10)

    # Treeview para Contenedores
    columnas_contenedores = ("Número de Contenedor", "Peso Neto (unidades)")
    tree_contenedores = ttk.Treeview(frame_contenedores, columns=columnas_contenedores, show="headings", height=5)
    tree_contenedores.heading("Número de Contenedor", text="Número de Contenedor")
    tree_contenedores.heading("Peso Neto (unidades)", text="Peso Neto (unidades)")

    tree_contenedores.column("Número de Contenedor", anchor='center', width=300)
    tree_contenedores.column("Peso Neto (unidades)", anchor='center', width=300)

    # Insertar datos de contenedores
    for i, contenedor in enumerate(contenedores_volumenes, start=1):
        tree_contenedores.insert("", "end", iid=i-1, values=(f"Contenedor {i}", f"{contenedor:,.2f}"))

    # Scrollbar para Contenedores
    scrollbar_contenedores = ttk.Scrollbar(frame_contenedores, orient="vertical", command=tree_contenedores.yview)
    tree_contenedores.configure(yscroll=scrollbar_contenedores.set)
    scrollbar_contenedores.pack(side='right', fill='y')
    tree_contenedores.pack(fill='both', expand=True)

    def on_select(event):
        selected_item = tree_contenedores.selection()
        if not selected_item:
            return

        contenedor_index = int(selected_item[0])

        # Crear una nueva ventana para mostrar los mensajes
        mensajes_ventana = tk.Toplevel(root)
        mensajes_ventana.title(f"Mensajes para Contenedor {contenedor_index + 1}")
        mensajes_ventana.geometry("800x400")  # Ajuste del tamaño de la ventana de mensajes

        # Configurar estilo para la nueva ventana
        style_mensajes = ttk.Style(mensajes_ventana)
        style_mensajes.theme_use("clam")
        style_mensajes.configure("Treeview.Heading", font=("Arial", 12, "bold"), background="#001955", foreground="white")
        style_mensajes.configure("Treeview", font=("Arial", 12), rowheight=30, fieldbackground="#f0f0f0")
        style_mensajes.map("Treeview", background=[("selected", "#ADD8E6")], foreground=[("selected", "black")])

        # Frame para los mensajes
        frame_mensajes = ttk.Frame(mensajes_ventana)
        frame_mensajes.pack(fill='both', expand=True)

        # Scrollbar vertical
        scrollbar_v = ttk.Scrollbar(frame_mensajes, orient='vertical')
        scrollbar_v.pack(side='right', fill='y')

        # Scrollbar horizontal
        scrollbar_h = ttk.Scrollbar(mensajes_ventana, orient='horizontal')
        scrollbar_h.pack(side='bottom', fill='x')

        # Definir las columnas que deseas mostrar, incluyendo 'Lote'
        columnas_mensajes = list(df_final.columns)
        if 'Lote' not in columnas_mensajes:
            columnas_mensajes.append('Lote')  # Añadir 'Lote' si no está presente

        # Crear el Treeview
        tree_mensajes = ttk.Treeview(frame_mensajes, columns=columnas_mensajes, show="headings",
                                     yscrollcommand=scrollbar_v.set, xscrollcommand=scrollbar_h.set)
        tree_mensajes.pack(side='left', fill='both', expand=True)

        scrollbar_v.config(command=tree_mensajes.yview)
        scrollbar_h.config(command=tree_mensajes.xview)

        # Definir encabezados y configurar las columnas
        for col in columnas_mensajes:
            # Usar el mapeo para los encabezados
            encabezado = COLUMN_MAPPING.get(col, col)
            tree_mensajes.heading(col, text=encabezado)
            tree_mensajes.column(col, width=100, minwidth=100, stretch=True)
            # Puedes ajustar 'width' y 'minwidth' según tus necesidades

        # Insertar datos de mensajes
        for mensaje_idx in mensajes_contenedores[contenedor_index]:
            mensaje = df_final.loc[mensaje_idx]
            valores = [mensaje[col] for col in columnas_mensajes]
            tree_mensajes.insert("", "end", values=valores)

    # Bind del evento de selección
    tree_contenedores.bind("<<TreeviewSelect>>", on_select)

    # Función para exportar y cerrar la ventana
    def exportar_y_cerrar():
        exportar_a_excel(contenedores_volumenes, mensajes_contenedores, df_final, columnas_mapeadas)
        root.destroy()  # Cerrar la ventana "Resultados de Contenedores"

    # Botón para Exportar a Excel
    boton_exportar = ttk.Button(root, text="Exportar a Excel", command=exportar_y_cerrar)
    boton_exportar.pack(pady=10)

    # Verificación de integridad antes de iniciar
    filas_asignadas = sum(len(mensajes) for mensajes in mensajes_contenedores)
    total_filas = df_final.shape[0]
    if filas_asignadas != total_filas:
        messagebox.showerror("Error de Integridad", f"Se han asignado {filas_asignadas} filas, pero el total es {total_filas}.")
        print(f"⚠️ Error: Se han asignado {filas_asignadas} filas, pero el total es {total_filas}.")
    else:
        print("✅ Todas las filas han sido asignadas correctamente a contenedores.")

    # No llamamos a root.mainloop() aquí, ya que el bucle principal ya está corriendo en la ventana principal
    # Si deseas que esta ventana sea modal, puedes usar grab_set()
    root.grab_set()


def main_proceso(df_final, columnas_mapeadas, capacidad_contenedor_max):
    try:
        # Verificar que las columnas requeridas estén presentes
        required_columns = ['Material', 'Texto de mensaje', 'Bruto', 'Neto', 'Volumen', 'Importe',
                            'LibrUtiliz', 'Contador', 'Cliente', 'Nombre', 'Doc.comer.', 'Grupo', 'Lote']
        for col in required_columns:
            if col not in df_final.columns:
                messagebox.showerror("Error", f"La columna '{col}' no está presente en los datos.")
                print(f"❌ Error: La columna '{col}' no está presente en los datos.")
                return

        # Rellenar valores nulos en la columna 'Nombre' si es necesario
        df_final['Nombre'] = df_final['Nombre'].fillna('Sin Nombre')

        # Calcular contenedores
        contenedores_volumenes, mensajes_contenedores = calcular_contenedores(df_final, columnas_mapeadas, capacidad_contenedor_max)

        # Calcular totales
        totales = calcular_totales(df_final)

        # Mostrar resultados
        mostrar_resultados(
            totales, contenedores_volumenes, mensajes_contenedores, df_final, columnas_mapeadas, capacidad_contenedor_max
        )
    except Exception as e:
        messagebox.showerror("Error", f"Error en el proceso principal: {e}")
        print(f"❌ Error en el proceso principal: {e}")

def cargar_datos(ruta_archivo):
    try:
        df = pd.read_excel(ruta_archivo)
        print(f"✅ Datos cargados exitosamente desde '{ruta_archivo}'.")
        return df
    except Exception as e:
        messagebox.showerror("Error de Carga", f"Ocurrió un error al cargar el archivo: {e}")
        print(f"❌ Error al cargar el archivo: {e}")
        return None

def main():
    try:
        # Crear la ventana principal para seleccionar el archivo de datos
        root = tk.Tk()
        root.withdraw()  # Ocultar la ventana principal

        messagebox.showinfo("Seleccionar Archivo", "Seleccione el archivo Excel que contiene los datos.")

        ruta_archivo = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Archivos de Excel", "*.xlsx *.xls")]
        )

        if not ruta_archivo:
            messagebox.showwarning("Sin Selección", "No se ha seleccionado ningún archivo. El programa se cerrará.")
            print("⚠️ No se ha seleccionado ningún archivo. El programa se cerrará.")
            return

        df_final = cargar_datos(ruta_archivo)
        if df_final is None:
            return

        # Definir las columnas mapeadas si es necesario (ejemplo)
        columnas_mapeadas = {
            'Bruto': 'Peso bruto (kg)',
            'Neto': 'Peso neto (kg)',
            'Volumen': 'Volumen (m3)',
            'Importe': 'Valor FOB',
            'Doc.comer.': 'Pedido',
            'LibrUtiliz': 'Cajas',
            'Total peso Bruto': 'Total peso Bruto (kg)',
            'Total peso Neto': 'Total peso Neto (kg)',
            'Total Volumen': 'Total volumen',
            'Total Importe': 'Ventas totales FOB',
            'Total LibrUtiliz': 'Total Cajas'
        }

        # Definir la capacidad máxima del contenedor (ejemplo)
        capacidad_contenedor_max = 1000  # Ajusta este valor según tus necesidades

        # Llamar a la función principal de procesamiento
        main_proceso(df_final, columnas_mapeadas, capacidad_contenedor_max)
    except Exception as e:
        messagebox.showerror("Error", f"Error en la función main: {e}")
        print(f"❌ Error en la función main: {e}")

if __name__ == "__main__":
    main()